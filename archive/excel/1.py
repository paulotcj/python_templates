from openpyxl import load_workbook

class AppComponentsExcel:
    local_filename = ""
    workbook = None
    metadata_sheet_name = ""
    appComponent_metadata = []
    appComponent_data = []
    
    #---------------------------------------------------------
    def __init__(self, fileName):
        local_filename = fileName
        
        #Read file
        self.workbook = load_workbook(filename=local_filename)
        # Check sheetnames
        # sheetnames = workbook.sheetnames
        # print(sheetnames)

    #---------------------------------------------------------
    def ReadMetadata(self, sheetName):
        self.metadata_sheet_name = sheetName
        #Read metadata
        sheet = self.workbook[self.metadata_sheet_name]

        for row_data in sheet.iter_rows(values_only=True, min_row=2):
            dict_row = {
                "table" : row_data[0],
                "order" : row_data[1],
                "column" : row_data[2],
                "column_type" : row_data[3],
                "data_type" : row_data[4],
                "annotations" : row_data[5],
            }
            self.appComponent_metadata.append(dict_row)

        return self.appComponent_metadata

    #end of def ReadMetadata(self, sheetName):    

    #---------------------------------------------------------   
    def ReadData(self, sheetName):
        #Read Data
        sheet = self.workbook[sheetName]

        for row_data in sheet.iter_rows(values_only=True, min_row=2, max_col=23):
            count = 0
            new_row = []
            if row_data[0] == None:
                continue
            #----
            for col in row_data:
                dict_row = {
                    "table" : self.appComponent_metadata[count]["table"],
                    "order" : self.appComponent_metadata[count]["order"],
                    "column" : self.appComponent_metadata[count]["column"],
                    "value" : col,
                    "column_type" : self.appComponent_metadata[count]["column_type"],
                    "data_type" : self.appComponent_metadata[count]["data_type"],
                    "annotations" : self.appComponent_metadata[count]["annotations"]
                }
                new_row.append(dict_row)
                count += 1
            #----
            self.appComponent_data.append(new_row)
        # end of for row_data in sheet

        return self.appComponent_data
    #end of def ReadData(self, sheetName):
    
    #---------------------------------------------------------

#end of class AppComponentsExcel
#---------------------------------------------------------
          


#---------------------------------------------------------    
class AppComponentsSQL:

    #---------------------------------------------------------    
    def __sql_value_tostring(self, value, data_type):
        new_value = None
        if value == None:
            new_value = "NULL"
        elif data_type == "string":
            temp = str( value ).replace("\n","")
            new_value = f'\'{  temp  }\''
        elif data_type == "bool":
            new_value = str( value ).upper()
        else:
            new_value = str( value )

        return new_value
    #---------------------------------------------------------    

    #---------------------------------------------------------    
    def Inserts(self, appComponent_data):
        #----
        columns_list = []
        for row in appComponent_data[0]:
            if row["column_type"] == "column":

                columns_list.append( row["column"] )

        temp = ', '.join(columns_list)
        columns = f'( { temp } )'

        #----
        row_list = []
        for row in appComponent_data:
            value_col_list = []
            for col in row:

                if col["column_type"] == "column":
                    value = self.__sql_value_tostring(col["value"] , col["data_type"])

                    value_col_list.append(value)
                # end of if col["column_type"] == "column":
            #end of for col in row:         

            temp = ", ".join(value_col_list)
            row_list.append( f"( {temp} )" )
        #end of for row in appComponent_data:
        
        insert_values = ",\n".join(row_list)

        #----
    
        #getting data from row[0] and column[0] -> then ["table"]
        table_name = appComponent_data[0][0]["table"]
        insert_header = f'INSERT INTO TABLE { table_name }\n{ columns }\nVALUES'
        inserts = f'{insert_header}\n{insert_values}\nON CONFLICT DO NOTHING;'

        print(inserts)


        return inserts

    #end of Inserts
    #---------------------------------------------------------

    #---------------------------------------------------------    
    def Updates(self, appComponent_data):

        table_name = appComponent_data[0][0]["table"]
        row_list = []
        for row in appComponent_data:
            where_list = []
            col_list = []
            for col in row:
                if col["column_type"] != "column":
                    continue
                
                #----
                value = self.__sql_value_tostring(col["value"], col["data_type"])
                temp_col = f'{col["column"]} = {value}'

                if str(col["annotations"]).lower() == "primarykey":
                    where_list.append(temp_col)
                else:
                    col_list.append(temp_col)
                #----
            #end of for col in row:

            temp_col = ", ".join(col_list)
            temp_where =  " AND ".join(where_list)
            temp_update =  f'UPDATE {table_name} SET {temp_col} WHERE {temp_where};'

            row_list.append(temp_update)
        #end of for row in appComponent_data:


        update = "\n".join(row_list)

        # print(update)

        return update
    #end of Updates(self, appComponent_data):
    #---------------------------------------------------------            

#end of AppComponentsSQL
#---------------------------------------------------------    


obj_excel = AppComponentsExcel('xxxxxxxxxxxxxxxxxxx')
obj_excel.ReadMetadata('AppComponents_metadata')
obj_excel.ReadData('Appcomponents_data')


dbscripts = AppComponentsSQL()


# -----------------------------------------------
#Build inserts
# inserts should allow conflict

# inserts = dbscripts.Inserts(obj_excel.appComponent_data)

# with open('appcomponent_inserts.sql', 'w') as file:
#     file.write(inserts)

# -----------------------------------------------
#Build updates
#  updates should be straight forward, and they are not supposed to fail

updates = dbscripts.Updates(obj_excel.appComponent_data)
with open('appcomponent_updates.sql', 'w') as file:
    file.write(updates)

# -----------------------------------------------
#Build migrations


quit()
